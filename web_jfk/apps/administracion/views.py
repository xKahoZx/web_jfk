#!/usr/bin/python
# -*- coding: utf-8 -*-
from django.shortcuts import render_to_response
from django.template import RequestContext
from web_jfk.apps.administracion.models import *
from django.http import HttpResponseRedirect
from django.contrib.auth.models import User
from datetime import date, datetime
import os
import openpyxl
from web_jfk.settings import MEDIA_ROOT

def consulta_institucion():	
	try:
		query = institucion.objects.get(id = 1)
	except:
		query = []
	return query
#CRUD noticias
#Crear nueva noticia

def add_noticia_view(request):
	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Noticias":
		if request.method == "POST":
			new_noticia = noticia()
			new_noticia.titulo = request.POST['titulo'].strip()
			new_noticia.copete = request.POST['subtitulo'].strip()
			new_noticia.cuerpo = request.POST['cuerpo'].strip()
			new_noticia.autor = request.POST['autor'].strip()
			new_noticia.imagen = request.FILES['imagen']
			new_noticia.fecha =  date.today()
			new_noticia.save()
			return HttpResponseRedirect('/noticias/page/1/')
			ctx = {'institucion':consulta_institucion()}
		return render_to_response('administracion/admin_noticia.html',ctx, context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect('/noticias/page/1/')
#editar noticia por id
def edit_noticia_view(request, id_noticia):

	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Noticias":
		noticia_id = noticia.objects.get(id = id_noticia)
		if request.method == "POST":
			noticia_id.titulo = request.POST['titulo'].strip()
			noticia_id.copete = request.POST['subtitulo'].strip()
			noticia_id.cuerpo = request.POST['cuerpo'].strip()
			noticia_id.autor = request.POST['autor'].strip()
			try:
				imagen = request.FILES['imagen']
				if os.path.exists(noticia_id.imagen._get_path()) and os.path.isfile(noticia_id.imagen._get_path()):
					os.remove(noticia_id.imagen._get_path())
				noticia_id.imagen = imagen
			except KeyError:
				pass
			noticia_id.save()
			return HttpResponseRedirect('/noticias/page/1')
		num = len(noticia_id.cuerpo)
		ctx = {'noticia': noticia_id,'num':num,'institucion':consulta_institucion()}
		return render_to_response('administracion/admin_noticia.html', ctx, context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect('/noticias/page/1/')
#eliminar noticia por id
def del_noticia_view(request, id_noticia):
	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Noticias":
		del_noticia = noticia.objects.get(id = id_noticia)
		if os.path.exists(del_noticia.imagen._get_path()) and os.path.isfile(del_noticia.imagen._get_path()):
			os.remove(del_noticia.imagen._get_path())
		del_noticia.delete()
		return HttpResponseRedirect('/noticias/page/1')
	else:
		return HttpResponseRedirect ('/noticias/page/1/')

#Album - Media
#Crear nuevo album
def add_album_view(request):
	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Multimedia":
		if request.method == "POST":
			new_album = album()
			new_album.titulo = request.POST['titulo'].strip()
			new_album.descripcion = request.POST['descripcion'].strip()

			if len(request.FILES.getlist('imagenes')) >= 6:
				new_album.save()
				for p in request.FILES.getlist('imagenes'):
					new_img = imagen_album()
					new_img.imagen = p
					new_img.album = new_album
					new_img.save() 
				return HttpResponseRedirect('/albums/page/1')
			else:
				caracteres = len(new_album.descripcion)
				ctx ={'titulo':new_album.titulo, 'desc':new_album.descripcion, 'ban':True, 'num':caracteres,'institucion':consulta_institucion()}
				return render_to_response('administracion/admin_album.html',ctx,  context_instance= RequestContext(request))
		return render_to_response('administracion/admin_album.html',  context_instance= RequestContext(request))
	else:
		return HttpResponseRedirect('/albums/page/1')
#editar album segun id
def edit_album_view(request, id_album):
	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Multimedia":
		album_id = album.objects.get(id = id_album)
		if request.method == "POST":
			album_id.titulo = request.POST['titulo'].strip()
			album_id.descripcion = request.POST['descripcion'].strip()
			for p in request.FILES.getlist('imagenes'):
				new_img = imagen_album()
				new_img.imagen = p
				new_img.album = album_id
				new_img.save() 
			album_id.save()
			return HttpResponseRedirect('/edit-album/%s' %album_id.id)
		caracteres = len(album_id.descripcion)
		num_img = len(imagen_album.objects.filter(album = album_id))
		ctx = {'album':album_id, 'num':caracteres,'num_img':num_img,'institucion':consulta_institucion()}
		return render_to_response('administracion/admin_album.html',ctx, context_instance=RequestContext(request))
	else:
		return HttpResponseRedirect('/albums/page/1')
#eliminar album segun id y sus imagenes asociadas
def del_album_view(request, id_album):
	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Multimedia":
		del_album = album.objects.get(id = id_album)
		imagenes = imagen_album.objects.filter(album = del_album)
		for del_img in imagenes:
			if os.path.exists(del_img.imagen._get_path()) and os.path.isfile(del_img.imagen._get_path()):
				os.remove(del_img.imagen._get_path())
			del_img.delete()
		del_album.delete()
		return HttpResponseRedirect('/albums/page/1')
	else:
		return HttpResponseRedirect('/albums/page/1')
#editar imagen agregada a un album segun id
def edit_img_view(request, id_img):
	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Multimedia":
		item = imagen_album.objects.get(id = id_img)
		if request.method=="POST":	
			img = request.FILES['imagen']
			if os.path.exists(item.imagen._get_path()) and os.path.isfile(item.imagen._get_path()):
					os.remove(item.imagen._get_path())
			item.imagen = img
			item.save()		
			id = album.objects.get(imagenes = id_img).id
			return HttpResponseRedirect('/edit-album/%s' %id)
		ctx = {'item':item,'institucion':consulta_institucion()}

		return render_to_response('administracion/edit_item.html',ctx, context_instance=RequestContext(request))
	else:
		return HttpResponseRedirect('/albums/page/1')
#eliminar imagen de un album por id
def del_img_album_view(request, id_img):

	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Multimedia":
		del_img = imagen_album.objects.get(id = id_img)
		id_album = album.objects.get(id = del_img.album.id)
		try:
			if len(imagen_album.objects.filter(album = id_album.id)) > 6:
				if os.path.exists(del_img.imagen._get_path()) and os.path.isfile(del_img.imagen._get_path()):
					os.remove(del_img.imagen._get_path())
				del_img.delete()
			return HttpResponseRedirect('/edit-album/%s' %id_album.id)
		except:
			return HttpResponseRedirect('/edit-album/%s' %id_album.id)
	else:
		return HttpResponseRedirect ('/albums/page/1')

#crear nueva imagen 
def add_item_view(request):
	if request.method == "POST":
		item = slider_show()
		item.imagen = request.FILES['imagen']
		item.save()
	return HttpResponseRedirect('/lista_items_carrousel')
#editar imagen
def edit_item_view(request, id_img):

	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Multimedia":
		item = slider_show.objects.get(id = id_img)
		if request.method=="POST":
			nueva_img = request.FILES['imagen']
			if os.path.exists(item.imagen._get_path()) and os.path.isfile(item.imagen._get_path()):
				os.remove(item.imagen._get_path())
			item.imagen = nueva_img
			item.save()
			return HttpResponseRedirect('/lista_items_carrousel')
		ctx = {'item':item,'institucion':consulta_institucion()}
		return render_to_response('administracion/edit_item.html',ctx, context_instance=RequestContext(request))
	else:
		return HttpResponseRedirect('/')

#modificar estado de la imagen de visible a no visible y vice versa
def modificar_estado_view(request, id_item, opc):

	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Multimedia":
		item = slider_show.objects.get(id = id_item)
		if opc == 'activar':
			item.estado = True
		else:
			item.estado = False
		item.save()
		return HttpResponseRedirect('/lista_items_carrousel')
	else:
		return HttpResponseRedirect('/')

#eliminar imagen
def del_item_view(request, id_item):

	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Multimedia":
		try:
			del_item = slider_show.objects.get(id = id_item)
			if os.path.exists(del_item.imagen._get_path()) and os.path.isfile(del_item.imagen._get_path()):
				os.remove(del_item.imagen._get_path())
			del_item.delete()
			return HttpResponseRedirect('/lista_items_carrousel')
		except:
			return HttpResponseRedirect('/lista_items_carrousel')
	else:
		return HttpResponseRedirect ('/')

#CRUD Eventos
#crear nuevo evento
def add_evento_view(request):
	
	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Noticias":
		if request.method == "POST":
			new_evento = calendario_eventos()
			new_evento.titulo = request.POST['titulo']
			new_evento.descripcion = request.POST['descripcion']
			new_evento.imagen = request.FILES['imagen']
			new_evento.fecha = request.POST['fecha']
			new_evento.anio = new_evento.fecha[0:4]
			new_evento.save()
			return HttpResponseRedirect('/eventos')
		ctx = {'institucion':consulta_institucion()}
		return render_to_response('administracion/administracion_eventos.html',ctx,context_instance = RequestContext(request))
	else:
		return  HttpResponseRedirect('/eventos')
		
#eliminar evento por id
def del_evento_view(request, id_evento):

	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Noticias":
		try:
			del_evento = calendario_eventos.objects.get(pk = id_evento)
			if os.path.exists(del_evento.imagen._get_path()) and os.path.isfile(del_evento.imagen._get_path()):
				os.remove(del_evento.imagen._get_path())
			del_evento.delete()
			return HttpResponseRedirect('/eventos')
		except:
			return HttpResponseRedirect('/eventos')
	else:
		return  HttpResponseRedirect('/eventos')

		
#editar evento por id
def edit_evento_view(request, id_evento):
	if request.user.is_superuser or  request.user.administrador.tipo_usuario == "Administrador Noticias":
		edit_evento = calendario_eventos.objects.get(id = id_evento)
		if request.method == "POST":
			edit_evento.titulo = request.POST['titulo']
			edit_evento.descripcion = request.POST['descripcion']
			edit_evento.fecha = request.POST['fecha']	
			edit_evento.anio = edit_evento.fecha[0:4]
			try:
				imagen = request.FILES['imagen']
				if os.path.exists(edit_evento.imagen._get_path()) and os.path.isfile(edit_evento.imagen._get_path()):
					os.remove(edit_evento.imagen._get_path())
				edit_evento.imagen = imagen
			except KeyError:
				pass
			fecha = date.today()
			if ( str(edit_evento.fecha) > str(fecha) or str(edit_evento.fecha) == str(fecha)):
				edit_evento.estado = True
			if( str(edit_evento.fecha) < str(fecha)):
				edit_evento.estado = False

			edit_evento.save()
			return HttpResponseRedirect('/eventos')
		ctx = {'evento': edit_evento,'institucion':consulta_institucion()}
		return render_to_response('administracion/administracion_eventos.html', ctx, context_instance = RequestContext(request))
	else:
		return  HttpResponseRedirect('/eventos')
#Verifica la fecha del evento si esta o no aun por suceder
def verificar_fecha_evento():
	eventos = calendario_eventos.objects.filter(estado = True)
	fecha = date.today()
	for p in eventos:
		if( str(fecha) > str(p.fecha)):
			p.estado = False
			p.save()
#CRUD OFERTA EDUCATIVA
#crear nueva oferta educativa
def add_oferta_view(request):
	if request.user.is_superuser:
		if request.method == "POST":
			oferta = oferta_educativa()
			oferta.fecha_apertura = request.POST['fecha_inicio']
			oferta.fecha_cierre = request.POST['fecha_fin']
			oferta.nivel = request.POST['nivel']
			oferta.grado = request.POST['grado']		
			oferta.cupos = request.POST['cupos']
			oferta.cupos_disponibles = oferta.cupos
			oferta.jornada = request.POST['jornada']		
			oferta.sede  = sede.objects.get(nombre = request.POST['sede'])
			if (str(oferta.fecha_apertura) > str(date.today())):
				oferta.estado = False
			oferta.save()
			return HttpResponseRedirect('/lista_ofertas')
		ctx = {'institucion':consulta_institucion()}
		return render_to_response('administracion/crear_oferta.html', ctx,context_instance=RequestContext(request))
	else:
		return  HttpResponseRedirect('/ofertas_educativas')
#editar oferta educativa por id
def edit_oferta_view(request, id_oferta):
	if  request.user.is_superuser:			
		oferta = oferta_educativa.objects.get(pk = id_oferta)
		if request.method == "POST":
			cupos = oferta.cupos	
			oferta.cupos = request.POST['cupos']
			oferta.fecha_apertura = request.POST['fecha_inicio']
			oferta.fecha_cierre = request.POST['fecha_fin']
			if cupos > oferta.cupos:
				cupos = cupos - oferta.cupos
				oferta.cupos_disponibles = oferta.cupos_disponibles - cupos
			else:
				cupos = int(oferta.cupos) - cupos	
				oferta.cupos_disponibles = oferta.cupos_disponibles + cupos

			if oferta.cupos_disponibles <= 0:
				oferta.estado = False
			else:
				oferta.estado = True
			oferta.save()
			return HttpResponseRedirect('/lista_ofertas')
		
		ctx = {'oferta': oferta,'institucion':consulta_institucion()}
		return render_to_response('administracion/editar_oferta.html', ctx, context_instance=RequestContext(request))
	else:
		return  HttpResponseRedirect('/ofertas_educativas')
#eliminar oferta educativa por id
def delete_oferta_view(request, id_oferta):
	if request.user.is_superuser:
		oferta = oferta_educativa.objects.get(id = id_oferta)
		inscripciones = inscripcion.objects.filter(oferta__id = id_oferta)
		for p in inscripciones:
			p.delete()
		oferta.delete()
		return HttpResponseRedirect('/lista_ofertas')
	else:
		return  HttpResponseRedirect('/ofertas_educativas')
#verifica si hay ofertas educativas para habilitar  o inhabilitar de acuerdo a la fecha.
def verificar_fecha():
	ofertas = oferta_educativa.objects.all()
	fecha = date.today()
	for p in ofertas:
		if( str(p.fecha_cierre) < str(fecha)):
			p.estado = False
			p.save()
		
		if( str(p.fecha_apertura) == str(fecha) or (str(p.fecha_apertura) < str(fecha)) and str(p.fecha_cierre) > str(fecha) ):
			if(p.cupos_disponibles > 0):
				p.estado = True
				p.save()
	
def inscripcion_view(request, id_oferta):
	if request.method == "POST":

		oferta_ed = oferta_educativa.objects.get(id = id_oferta)
		if oferta_ed.cupos_disponibles > 0:
			registro = inscripcion()
			registro.nombres_alumno = request.POST['nombres_a']
			registro.apellidos_alumno = request.POST['apellidos_a']
			registro.nombres = request.POST['nombres']
			registro.apellidos = request.POST['apellidos']
			registro.identificacion = request.POST['identificacion']
			registro.telefono_1 = request.POST['telefono']
			registro.telefono_2 = request.POST['celular']
			registro.edad = request.POST['edad']	
			registro.fecha_inscripcion = date.today()
			oferta_ed.cupos_disponibles = oferta_ed.cupos_disponibles - 1
			if oferta_ed.cupos_disponibles == 0:	
				oferta_ed.estado = False
			oferta_ed.save()
			registro.oferta = oferta_ed
			registro.save()		
	return HttpResponseRedirect('/ofertas_educativas')

#CRUD Documentos academicos
#crear nuevo documento
def add_documento_view(request):
	if request.method == "POST":
		new_documento = documento()
		new_documento.titulo = request.POST['titulo']
		new_documento.descripcion = request.POST['descripcion']
		try:
			if request.user.docente.jornada == "Tarde" and request.user.docente.sede.nombre == "John F. Kennedy":
 				new_documento.materia = request.POST['materia']
				new_documento.grado = request.POST['grado']
		except:
			pass
		new_documento.fecha = date.today()
		new_documento.autor =  request.user
		new_documento.documento = request.FILES['documento']			
		new_documento.save()
		return HttpResponseRedirect('/aula_virtual')	
	ctx = {'institucion':consulta_institucion()}
	return render_to_response('administracion/admin_documento.html', ctx,context_instance = RequestContext(request))
def edit_documento_view(request, id_documento):

	edit_documento = documento.objects.get(id = id_documento)
	if request.user == edit_documento.autor:
		if request.method == "POST":
			edit_documento.titulo = request.POST['titulo']
			edit_documento.descripcion = request.POST['descripcion']
			try:
				if request.user.docente.jornada == "Tarde" and request.user.docente.sede.nombre == "John F. Kennedy":
	 				edit_documento.materia = request.POST['materia']
					edit_documento.grado = request.POST['grado']
			except:
				pass
			try:
				new_documento = request.FILES['documento']
				if os.path.exists(edit_documento.documento._get_path()) and os.path.isfile(edit_documento.documento._get_path()):
					os.remove(edit_documento.documento._get_path())
				edit_documento.documento = new_documento
			except KeyError:
				pass
			edit_documento.save()
			return HttpResponseRedirect('/mis_documentos')
		ctx = {'documento':edit_documento ,'institucion':consulta_institucion()}
		return render_to_response('administracion/admin_documento.html', ctx, context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect('/mis_documentos')
def del_documento_view(request, id_documento):
	
	del_documento = documento.objects.get(id = id_documento)
	if request.user == del_documento.autor:
		if os.path.exists(del_documento.documento._get_path()) and os.path.isfile(del_documento.documento._get_path()):
			os.remove(del_documento.documento._get_path())
		del_documento.delete()
	return HttpResponseRedirect('/mis_documentos')

#CRUD usuarios

def add_usuario_view(request):

	if request.user.is_superuser:
		sedes = sede.objects.all()
		if request.method == "POST":
			if (request.POST['tipo_select']== "Docente"): 
				new_usuario = docente()
				new_usuario.foto = request.FILES['foto']
			else:
				new_usuario = administrador()
			
			new_usuario.nombres = request.POST['nombres']
			new_usuario.apellidos = request.POST['apellidos']
			new_usuario.correo = request.POST['correo']
			try:
				u = User.objects.create_user(username = request.POST['username'], password = request.POST['password'])
				u.save()
			except:
				men = "El nombre de usuario ya se encuentra registrado"
				ctx = {'sedes' : sedes, 'men':men,'institucion':consulta_institucion()}
				return render_to_response('administracion/crear_usuario.html', ctx, context_instance=RequestContext(request))
			if(request.POST['tipo_select'] == "Docente"):

				new_usuario.jornada = request.POST['jornada_select']
				if(request.POST['jornada_select'] == "Tarde" and request.POST['sede_select'] == "John F. Kennedy"):
					new_usuario.curso = request.POST['curso_select_b']
				else:
					new_usuario.curso = request.POST['curso_select_c']
				new_usuario.user = u	
				new_usuario.sede = sede.objects.get(nombre = request.POST['sede_select'])
				new_usuario.save()
				
			else:
				new_usuario.tipo_usuario = request.POST['rol_select']
				if request.POST['rol_select'] == "Administrador General":
					u.is_superuser = True
					u.is_staff = True
					u.save()
				new_usuario.user = u
				new_usuario.save()

			return HttpResponseRedirect('/administracion')
		ctx = {'sedes' : sedes,'institucion':consulta_institucion()}

		return render_to_response('administracion/crear_usuario.html', ctx, context_instance=RequestContext(request))
	else:
		return HttpResponseRedirect('/administracion')
def edit_usuario_view(request, tipo_usuario, id_user):
	if request.user.is_superuser:
		if tipo_usuario == "Docente":
			try:
				usuario = docente.objects.get(id = id_user)
			except:
				return HttpResponseRedirect('/administracion')
		else:
			if tipo_usuario == "Administrador":
				try:
					usuario = administrador.objects.get(id = id_user)
				except:
					return HttpResponseRedirect('/administracion')
			else:
				return HttpResponseRedirect('/administracion')
		if request.method == "POST":
			usuario.nombres = request.POST['nombres']
			usuario.apellidos = request.POST['apellidos']
			usuario.correo = request.POST['correo']
			try:
				if usuario.tipo == "Docente":
					usuario.jornada = request.POST['jornada_select']
					if(request.POST['jornada_select'] == "Tarde" and request.POST['sede_select'] == "John F. Kennedy"):
						usuario.curso = request.POST['curso_select_b']
					else:
						usuario.curso = request.POST['curso_select_c']
					usuario.sede = sede.objects.get(nombre = request.POST['sede_select'])
					try:
						usuario.foto = request.FILES['foto']
					except KeyError:
						pass
			except:
				usuario.tipo_usuario = request.POST['rol_select']
				u = usuario.user
				if request.POST['rol_select'] == "Administrador General":
					u.is_superuser = True
					u.is_staff = True
				else:
					u.is_staff = False
					u.is_superuser = False
				u.save()
			usuario.save()
			return HttpResponseRedirect('/administracion')
		sedes = sede.objects.all()
		ctx = {'user':usuario,'sedes':sedes,'institucion':consulta_institucion()}
		return  render_to_response('administracion/edit_usuario.html', ctx, context_instance = RequestContext(request))
	HttpResponseRedirect('/administracion')

def add_alumno_view(request):
	if request.method == "POST":
		try:
			query_user = User.objects.get(username = request.POST['username'])
			men = "El nombre de usuario ya se encuentra registrado"
			ctx = {'men':men,'institucion':consulta_institucion() }
			return render_to_response('administracion/registro_estudiante.html', ctx, context_instance=RequestContext(request))	
		except:
			pass
		identificacion = request.POST['identificacion']
		try:
			men = "El numero de identificacion ingresado ya se encuentra registrado"
			query_studiante = estudiante.objects.get(identificacion = request.POST['identificacion'])
			ctx = {'men':men,'institucion':consulta_institucion() }
			return render_to_response('administracion/registro_estudiante.html', ctx, context_instance=RequestContext(request))	
		except:
			pass
		#abrir el archivo excel
		document = openpyxl.load_workbook('%s/MultimediaData/doc_estudiantes/estudiantes.xlsx' % MEDIA_ROOT)
		#obtiene las hojas dentro del documento
		hojas = document.get_sheet_names()
		print hojas
		for p in hojas:
			#accede a la hoja indicada segun el ciclo
			hoja = document.get_sheet_by_name(p)
			print hoja
			#recorre la columna 1 y la fila es una variable que aumenta 
			row = 1
			#el ciclo for se cierra cuando encuentra un campo con valor None osea vacio
			for valor in range(300):
				num_identificacion = hoja.cell(row = row, column = 1).value
				print str(valor) + "ID BD " + str(num_identificacion) + "ID EN " + identificacion
				#si encuentra que las identificacion son iguales entra a la siguiente condicion
				if num_identificacion == long(identificacion):
					u = User.objects.create_user(username = request.POST['username'], password=request.POST['password'])
					u.save()
					new_estudiante = estudiante()	
					new_estudiante.user = u
					new_estudiante.correo = request.POST['email']
					nom_1 = hoja.cell(row = row, column = 2).value
					nom_2 = hoja.cell(row = row, column = 3).value
					ape_1 = hoja.cell(row = row, column = 4).value
					ape_2 = hoja.cell(row = row, column = 5).value
					try:
						new_estudiante.nombres = nom_1 + " " + nom_2
					except:
						new_estudiante.nombres = nom_1
					try:
						new_estudiante.apellidos = ape_1 + " " + ape_2
					except :
						new_estudiante.apellidos = ape_1
					new_estudiante.identificacion = num_identificacion
					new_estudiante.grado = hoja.cell(row = row, column = 6).value
					new_estudiante.save()
					hoja.cell(row = row  , column = 7).value = True	
					document.save('%s/MultimediaData/doc_estudiantes/estudiantes.xlsx' % MEDIA_ROOT)
					if request.user.is_superuser:
						return HttpResponseRedirect('/lista_estudiantes')
					return HttpResponseRedirect('/login')
				if hoja.cell(row = row + 1 , column = 1).value != None:
					row = row + 1
				else:
					break
		men = "El numero de identificacion ingresado no se encuentra registrado en nuestra base de datos."
		ctx = {'men':men ,'institucion':consulta_institucion()}
		return render_to_response('administracion/registro_estudiante.html', ctx, context_instance=RequestContext(request))	
					
	return render_to_response('administracion/registro_estudiante.html', context_instance=RequestContext(request))	
#Cambiar password
def edit_password_view(request, id_user):
	if request.user.is_superuser:
		if request.method == "POST":
			user = User.objects.get(id = id_user)
			user.set_password(request.POST['password'])
			user.save()
			return HttpResponseRedirect('/administracion')
		return render_to_response('administracion/cambiar_password.html', context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect('/administracion')
#eliminar usuario
def del_user_view(request, id_user):
	if request.user.is_superuser:
		del_user = User.objects.get(id = id_user)
		del_user.delete()
	return HttpResponseRedirect('/lista_estudiantes')

#SEDE
def add_sede_view(request):
	if request.user.is_superuser:
		try:
			query = institucion.objects.get(id = 1)
		except :
			return HttpResponseRedirect('/crear_institucion')
		if request.method == "POST":
			new_sede = sede()
			new_sede.nombre = request.POST['nombre']
			new_sede.nit 	= request.POST['nit']
			new_sede.telefono_1 = request.POST['telefono_1']
			new_sede.telefono_2 = request.POST['telefono_2']
			new_sede.direccion = request.POST['direccion']
			new_sede.correo = request.POST['correo']
			new_sede.institucion = institucion.objects.get(id = 1)
			new_sede.save()
			return HttpResponseRedirect('/add_coordinador/%s' % new_sede.id)
		ctx = {'institucion':consulta_institucion()}
		return render_to_response('administracion/admin_sede.html', ctx, context_instance = RequestContext(request))
	else:	
		return HttpResponseRedirect('/administracion')

def edit_sede_view(request, id_sede):
	if request.user.is_superuser:
		query_sede = sede.objects.get(id = id_sede)
		if request.method == "POST":
			query_sede.nombre = request.POST['nombre']
			query_sede.nit 	= request.POST['nit']
			query_sede.telefono_1 = request.POST['telefono_1']
			query_sede.telefono_2 = request.POST['telefono_2']
			query_sede.direccion = request.POST['direccion']
			query_sede.correo = request.POST['correo']
			query_sede.save()
			return HttpResponseRedirect('/ver_sede/%s' % query_sede.id)
		ctx = {'sede':query_sede,'institucion':consulta_institucion()}
		return render_to_response('administracion/admin_sede.html',ctx, context_instance = RequestContext(request))
	else:	
		return HttpResponseRedirect('/administracion')
#INSTITUCION
def edit_funcionario_view(request, id_user):
	if request.user.is_superuser:
		query_funcionario = funcionario.objects.get(id = id_user)
		if request.method=="POST":
			query_funcionario.nombres = request.POST['nombres']
			query_funcionario.apellidos = request.POST['apellidos']
			try:
				foto = request.FILES['foto']
				if os.path.exists(query_funcionario.foto._get_path()) and os.path.isfile(query_funcionario.foto._get_path()):
					os.remove(query_funcionario.foto._get_path())
				query_funcionario.foto = foto
			except KeyError:
				pass
			query_funcionario.correo = request.POST['correo']
			query_funcionario.save()
			try:
				id = sede.objects.get(coordinadores = query_funcionario).id
				return HttpResponseRedirect('/ver_sede/%s' % id )
			except:
				return HttpResponseRedirect('/institucion')
			return HttpResponseRedirect('/lista_sedes')
		bandera = "edit"
		ctx ={'funcionario':query_funcionario, 'bandera':bandera,'institucion':consulta_institucion()}
		return render_to_response('administracion/admin_funcionario.html', ctx,context_instance=RequestContext(request))
	else:
		return HttpResponseRedirect('/administracion')

def add_coordinador_view(request, id_sede):

	if request.user.is_superuser:
		query_sede = sede.objects.get(id = id_sede)
		if query_sede.coordinadores.count() < 2:
	 		jornada = ""
			if request.method == "POST":
				new_funcionario = funcionario()
				new_funcionario.nombres = request.POST['nombres']
				new_funcionario.apellidos = request.POST['apellidos']
				new_funcionario.correo = request.POST['correo']
				new_funcionario.foto = request.FILES['foto']
				new_funcionario.tipo_funcionario = "Coordinador"
				u = User.objects.create_user(username = request.POST['username'], password=request.POST['password'])
				u.save()
				new_funcionario.user = u 
				if query_sede.coordinadores.count() == 0:
					new_funcionario.jornada = "Mañana"
				else:
					new_funcionario.jornada = "Tarde"
				new_funcionario.save()
				query_sede.coordinadores.add(new_funcionario)
				query_sede.save()
				if query_sede.coordinadores.count() == 2:
					return HttpResponseRedirect('/lista_sedes')
			if query_sede.coordinadores.count() == 0:
				jornada = "mañana"
			else:
				jornada = "tarde"
			bandera = "crear"
			ctx = {'jornada':jornada ,'sede':query_sede.nombre, 'bandera':bandera,'institucion':consulta_institucion()}
			return render_to_response('administracion/admin_funcionario.html', ctx,context_instance=RequestContext(request))
		return HttpResponseRedirect('/lista_sedes')
	else:	
		return HttpResponseRedirect('/administracion')

def add_institucion_view(request):
	if request.user.is_superuser:
		if request.method == "POST":
			new_funcionario = funcionario()
			new_funcionario.nombres = request.POST['nombres']
			new_funcionario.apellidos = request.POST['apellidos']
			new_funcionario.correo = request.POST['correo']
			new_funcionario.foto = request.FILES['foto']
			new_funcionario.tipo_funcionario = "Rector"
			u = User.objects.create_user(username = request.POST['username'], password=request.POST['password'])
			u.is_staff = True
			u.is_superuser = True			
			u.save()
			new_funcionario.user = u
			new_funcionario.save()
			new_insti = institucion()
			new_insti.nombre = request.POST['nombre']
			new_insti.estudiantes = request.FILES['estudiantes']
			new_insti.escudo = request.FILES['escudo']
			new_insti.rector = new_funcionario
			new_insti.save()
			return HttpResponseRedirect('/institucion')
		ctx = {'institucion':consulta_institucion()}
		return render_to_response('administracion/crear_institucion.html',ctx,context_instance=RequestContext(request))
	else:
		return HttpResponseRedirect('/administracion')
def edit_institucion_view(request):
	if request.user.is_superuser:
		

		query_institucion = institucion.objects.get(id = 1)
		if request.method == "POST":
			query_institucion.nombre = request.POST['nombre']
			try:
				escudo = request.FILES['escudo']
				if os.path.exists(query_institucion.escudo._get_path()) and os.path.isfile(query_institucion.escudo._get_path()):
					os.remove(query_institucion.escudo._get_path())
				query_institucion.escudo = escudo
			except KeyError:
				pass
			try:
				estudiantes = request.FILES['estudiantes']
				if os.path.exists(query_institucion.estudiantes._get_path()) and os.path.isfile(query_institucion.estudiantes._get_path()):
					os.remove(query_institucion.estudiantes._get_path())
				query_institucion.estudiantes = estudiantes
				query_institucion.save()
				#eliminar_usuarios()
			except KeyError:
				pass
			query_institucion.save()
			return HttpResponseRedirect('/institucion')
		ctx = {'institucion':query_institucion}
		return render_to_response('administracion/edit_institucion.html', ctx, context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect('/administracion')


def eliminar_usuarios():

	document = openpyxl.load_workbook('%s/MultimediaData/doc_estudiantes/estudiantes.xlsx' % MEDIA_ROOT)
	#acceder a una hoja del documento
	hojas = document.get_sheet_names()

	estudiantes = estudiante.objects.all()
	for p in estudiantes:
		p.estado = False
		p.save()
	
	for p in hojas:
		hoja = document.get_sheet_by_name(p)
		row = 1
		for valor in range(50):
			if hoja.cell(row = row  , column = 4).value == None:				
				num_identificacion = hoja.cell(row = row, column = 1).value
				print num_identificacion
				try:
					verificar_estudiante = estudiante.objects.get(identificacion = num_identificacion)
					verificar_estudiante.estado = True
					verificar_estudiante.save()
					hoja.cell(row = row  , column = 4).value = True				
				except:
					hoja.cell(row = row  , column = 4).value = False				
				
						
			if hoja.cell(row = row + 1 , column = 1).value == None:
				break
			else:
				row = row + 1
	for p in estudiante.objects.filter(estado = False):
		p.user.delete()
		p.delete()
	document.save('%s/MultimediaData/doc_estudiantes/estudiantes.xlsx' % MEDIA_ROOT)



